from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Producto, Categoria, Subcategoria, MovimientoStock, Venta, DetalleVenta
from django.http import JsonResponse
from django.db.models import Q
from decimal import Decimal
from registration.decorators import rol_requerido, solo_administrador, solo_vendedor_o_admin
#imports para excel
import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


@solo_vendedor_o_admin
def lista_productos(request):
    """Vista principal del inventario - Solo vendedores y admin"""
    productos = Producto.objects.filter(activo=True).select_related('categoria', 'subcategoria')
    categorias = Categoria.objects.filter(activo=True)
    subcategorias = Subcategoria.objects.filter(activo=True)
    
    # Filtros
    categoria_filtro = request.GET.get('categoria')
    subcategoria_filtro = request.GET.get('subcategoria')
    busqueda = request.GET.get('busqueda')
    
    if categoria_filtro:
        productos = productos.filter(categoria_id=categoria_filtro)
    
    if subcategoria_filtro:
        productos = productos.filter(subcategoria_id=subcategoria_filtro)
    
    if busqueda:
        productos = productos.filter(nombre__icontains=busqueda) | productos.filter(codigo_sku__icontains=busqueda)
    
    # Estadisticas
    total_productos = productos.count()
    total_unidades = sum(p.stock for p in productos)
    
    # Contar productos con stock bajo y critico
    stock_bajo = sum(1 for p in productos if p.get_estado_stock() == 'BAJO')
    stock_critico = sum(1 for p in productos if p.get_estado_stock() == 'CRITICO')
    
    context = {
        'productos': productos,
        'categorias': categorias,
        'subcategorias': subcategorias,
        'total_productos': total_productos,
        'total_unidades': total_unidades,
        'stock_bajo': stock_bajo,
        'stock_critico': stock_critico,
    }
 
    return render(request, 'inventario/lista_productos.html', context)


@solo_vendedor_o_admin
def crear_producto(request):
    """Crear nuevo producto - Solo vendedores y admin"""
    
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre', '').strip()
            precio = request.POST.get('precio')
            stock = request.POST.get('stock')
            stock_minimo = request.POST.get('stock_minimo', 5)
            stock_critico = request.POST.get('stock_critico', 2)
            categoria_id = request.POST.get('categoria')
            
            if not nombre:
                messages.error(request, 'El nombre del producto es obligatorio')
                return redirect('inventario:lista_productos')
            
            if not categoria_id:
                messages.error(request, 'Debe seleccionar una categoría')
                return redirect('inventario:lista_productos')
            
            try:
                precio = float(precio)
                if precio <= 0:
                    messages.error(request, 'El precio debe ser mayor a 0')
                    return redirect('inventario:lista_productos')
            except (ValueError, TypeError):
                messages.error(request, 'El precio debe ser un número válido')
                return redirect('inventario:lista_productos')
            
            try:
                stock = int(stock)
                if stock < 0:
                    messages.error(request, 'El stock no puede ser negativo')
                    return redirect('inventario:lista_productos')
            except (ValueError, TypeError):
                messages.error(request, 'El stock debe ser un número válido')
                return redirect('inventario:lista_productos')
            
            try:
                stock_minimo = int(stock_minimo)
                stock_critico = int(stock_critico)
                
                if stock_minimo < 0 or stock_critico < 0:
                    messages.error(request, 'Los valores de stock mínimo y crítico no pueden ser negativos')
                    return redirect('inventario:lista_productos')
                
                if stock_critico > stock_minimo:
                    messages.error(request, 'El stock crítico no puede ser mayor al stock mínimo')
                    return redirect('inventario:lista_productos')
                    
            except (ValueError, TypeError):
                messages.error(request, 'Los valores de stock mínimo y crítico deben ser números válidos')
                return redirect('inventario:lista_productos')
            
            try:
                categoria = Categoria.objects.get(id=categoria_id, activo=True)
            except Categoria.DoesNotExist:
                messages.error(request, 'La categoría seleccionada no es válida')
                return redirect('inventario:lista_productos')
            
            subcategoria_id = request.POST.get('subcategoria')
            if subcategoria_id:
                try:
                    subcategoria = Subcategoria.objects.get(id=subcategoria_id, activo=True)
                except Subcategoria.DoesNotExist:
                    messages.error(request, 'La subcategoría seleccionada no es válida')
                    return redirect('inventario:lista_productos')
            else:
                subcategoria = None
            
            producto = Producto(
                nombre=nombre,
                categoria=categoria,
                subcategoria=subcategoria,
                descripcion=request.POST.get('descripcion', '').strip(),
                precio=precio,
                stock=stock,
                stock_minimo=stock_minimo,
                stock_critico=stock_critico,
            )
            
            if 'imagen' in request.FILES:
                producto.imagen = request.FILES['imagen']
            
            producto.save()
            messages.success(request, f'✓ Producto {producto.codigo_sku} creado exitosamente')
            return redirect('inventario:lista_productos')
            
        except Exception as e:
            messages.error(request, f'Error al crear producto: {str(e)}')
            return redirect('inventario:lista_productos')
    
    return redirect('inventario:lista_productos')


@solo_administrador
def editar_producto(request, pk):
    """Editar producto existente - SOLO ADMINISTRADOR"""
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre', '').strip()
            precio = request.POST.get('precio')
            stock = request.POST.get('stock')
            stock_minimo = request.POST.get('stock_minimo', 5)
            stock_critico = request.POST.get('stock_critico', 2)
            categoria_id = request.POST.get('categoria')
            
            if not nombre:
                messages.error(request, 'El nombre del producto es obligatorio')
                return redirect('inventario:lista_productos')
            
            if not categoria_id:
                messages.error(request, 'Debe seleccionar una categoría')
                return redirect('inventario:lista_productos')
            
            try:
                precio = float(precio)
                if precio <= 0:
                    messages.error(request, 'El precio debe ser mayor a 0')
                    return redirect('inventario:lista_productos')
            except (ValueError, TypeError):
                messages.error(request, 'El precio debe ser un número válido')
                return redirect('inventario:lista_productos')
            
            try:
                stock = int(stock)
                if stock < 0:
                    messages.error(request, 'El stock no puede ser negativo')
                    return redirect('inventario:lista_productos')
            except (ValueError, TypeError):
                messages.error(request, 'El stock debe ser un número válido')
                return redirect('inventario:lista_productos')

            try:
                stock_minimo = int(stock_minimo)
                stock_critico = int(stock_critico)
                
                if stock_minimo < 0 or stock_critico < 0:
                    messages.error(request, 'Los valores de stock mínimo y crítico no pueden ser negativos')
                    return redirect('inventario:lista_productos')
                
                if stock_critico > stock_minimo:
                    messages.error(request, 'El stock crítico no puede ser mayor al stock mínimo')
                    return redirect('inventario:lista_productos')
                    
            except (ValueError, TypeError):
                messages.error(request, 'Los valores de stock mínimo y crítico deben ser números válidos')
                return redirect('inventario:lista_productos')
            
            try:
                categoria = Categoria.objects.get(id=categoria_id, activo=True)
            except Categoria.DoesNotExist:
                messages.error(request, 'La categoría seleccionada no es válida')
                return redirect('inventario:lista_productos')
            
            subcategoria_id = request.POST.get('subcategoria')
            if subcategoria_id:
                try:
                    subcategoria = Subcategoria.objects.get(id=subcategoria_id, activo=True)
                except Subcategoria.DoesNotExist:
                    messages.error(request, 'La subcategoría seleccionada no es válida')
                    return redirect('inventario:lista_productos')
            else:
                subcategoria = None
            
            producto.nombre = nombre
            producto.categoria = categoria
            producto.subcategoria = subcategoria
            producto.descripcion = request.POST.get('descripcion', '').strip()
            producto.precio = precio
            producto.stock = stock
            producto.stock_minimo = stock_minimo
            producto.stock_critico = stock_critico
            
            if 'imagen' in request.FILES:
                producto.imagen = request.FILES['imagen']
            
            producto.save()
            
            messages.success(request, f'✓ Producto {producto.codigo_sku} actualizado exitosamente')
            return redirect('inventario:lista_productos')
            
        except Exception as e:
            messages.error(request, f'Error al actualizar producto: {str(e)}')
            return redirect('inventario:lista_productos')
    
    return redirect('inventario:lista_productos')


@solo_administrador
def eliminar_producto(request, pk):
    """Baja logica de producto - SOLO ADMINISTRADOR"""
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        producto.activo = False
        producto.save()
        messages.success(request, f'Producto {producto.codigo_sku} dado de baja exitosamente')
    
    return redirect('inventario:lista_productos')


@solo_vendedor_o_admin
def ajustar_stock(request, pk):
    """Vista para ajustar stock de un producto con historial - Vendedores y Admin"""
    producto = get_object_or_404(Producto, pk=pk)
    
    # Obtener perfil del usuario
    perfil = request.user.perfilusuario
    es_vendedor = perfil.rol.nombre == 'Vendedor'
    
    #Obtener historial de movimientos del producto
    movimientos = MovimientoStock.objects.filter(producto=producto).order_by('-fecha_movimiento')[:20]
    
    if request.method == 'POST':
        try:
            tipo_movimiento = request.POST.get('tipo_movimiento')
            cantidad = int(request.POST.get('cantidad'))
            motivo = request.POST.get('motivo')
            observaciones = request.POST.get('observaciones')
            
            # RESTRICCIÓN: Vendedor NO puede usar AJUSTE
            if es_vendedor and tipo_movimiento == 'AJUSTE':
                messages.error(request, '⛔ Los vendedores no pueden usar la función AJUSTE. Solo ENTRADA o SALIDA.')
                return redirect('inventario:ajustar_stock', pk=pk)
            
            #Validar Tipo de movimiento
            if not tipo_movimiento:
                messages.error(request, 'Debe seleccionar un tipo de movimiento')
                return redirect('inventario:ajustar_stock', pk=pk)
            
            #Validar la cantidad
            if cantidad <= 0:
                messages.error (request, 'La cantidad debe ser mayor a 0')
                return redirect('inventario:ajustar_stock', pk=pk)
            
            #Guardar stock anterior
            stock_anterior = producto.stock
            
            #Calcular nuevo stock segun el tipo de movimiento
            if tipo_movimiento == 'ENTRADA':
                producto.stock += cantidad
                
            elif tipo_movimiento == 'SALIDA':
                
                if producto.stock < cantidad:
                    messages.error(request, f'Stock insuficiente. Stock actual: {producto.stock}')
                    return redirect('inventario:ajustar_stock', pk=pk)
                producto.stock -= cantidad
                
            elif tipo_movimiento == 'AJUSTE':
                producto.stock = cantidad
                
            stock_nuevo = producto.stock
            producto.save()
            
            #Registrar producto en el historial
            MovimientoStock.objects.create(
                producto = producto,
                tipo = tipo_movimiento,
                cantidad = cantidad if tipo_movimiento != 'AJUSTE' else abs(stock_nuevo - stock_anterior),
                stock_anterior = stock_anterior,
                stock_nuevo = stock_nuevo,
                motivo = motivo,
                observaciones = observaciones,
                usuario = request.user if request.user.is_authenticated else None
            )
            
            messages.success(request, f'Stock actualizado correctamente, Nuevo Stock: {stock_nuevo}')
            return redirect('inventario:ajustar_stock', pk=pk)
        
        except Exception as e:
            messages.error(request, f'Error al ajustar stock: {str(e)}')
            return redirect('inventario:ajustar_stock', pk=pk)
        
    context = {
        'producto' : producto,
        'movimientos' : movimientos,
        'es_vendedor': es_vendedor,
    }
    
    return render(request, 'inventario/ajustar_stock.html', context)

@solo_vendedor_o_admin
def importar_productos(request):
    """Vista para importar productos desde Excel/CSV - Vendedores y Admin"""
    if request.method == 'GET':
        # Solo limpiar errores si viene con el parámetro 'limpiar' o si no hay errores previos
        if request.GET.get('limpiar') == '1' or not request.session.get('errores_importacion'):
            if 'errores_importacion' in request.session:
                del request.session['errores_importacion']
                request.session.modified = True
        
        return render(request, 'inventario/importar_productos.html')
    
    elif request.method == 'POST':
        # ... (mantén todo el código existente de importación)
        if 'archivo' not in request.FILES:
            messages.error(request, 'No se ha seleccionado ningun archivo')
            return redirect('inventario:importar_productos')
        
        archivo = request.FILES['archivo']
        
        if not archivo.name.endswith(('.xlsx', '.xls', '.csv')):
            messages.error(request, 'Formato de archivo no valido. Use .xlsx, .xls o .csv')
            return redirect('inventario:importar_productos')
        
        try:
            if archivo.name.endswith('.csv'):
                df = pd.read_csv(archivo)
            else:
                df = pd.read_excel(archivo)
                
            columnas_requeridas = ['codigo_sku', 'nombre', 'categoria', 'precio']
            columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
            
            if columnas_faltantes:
                messages.error(request, f'Faltan columnas requeridas: {", ".join(columnas_faltantes)}')
                return redirect('inventario:importar_productos')
            
            productos_validos = []
            productos_actualizados = 0
            errores = []
            
            for index, row in df.iterrows():
                fila_num = index + 2
                
                try:
                    if pd.isna(row['codigo_sku']):
                        errores.append(f'Fila {fila_num}: Falta el código SKU del producto')
                        continue
                    
                    if pd.isna(row['nombre']):
                        errores.append(f'Fila {fila_num}: Falta el nombre del producto')
                        continue
                        
                    if pd.isna(row['precio']):
                        errores.append(f'Fila {fila_num}: Falta el precio del producto')
                        continue
                    
                    try:
                        precio = float(row['precio'])
                        if precio <= 0:
                            errores.append(f'Fila {fila_num}: El precio debe ser mayor a 0')
                            continue
                    except:
                        errores.append(f'Fila {fila_num}: El precio tiene un formato inválido')
                        continue
                    
                    if pd.isna(row['categoria']):
                        errores.append(f'Fila {fila_num}: Falta la categoría del producto')
                        continue
                        
                    try:
                        categoria = Categoria.objects.get(nombre=row['categoria'])
                    except Categoria.DoesNotExist:
                        errores.append(f"Fila {fila_num}: La categoría '{row['categoria']}' no existe en el sistema. Créela primero en el admin.")
                        continue
                    
                    subcategoria = None
                    if not pd.isna(row.get('subcategoria')):
                        try:
                            subcategoria = Subcategoria.objects.get(nombre=row['subcategoria'])
                        except Subcategoria.DoesNotExist:
                            errores.append(f"Fila {fila_num}: La subcategoría '{row['subcategoria']}' no existe en el sistema. Créela primero en el admin.")
                            continue
                    
                    producto_existente = Producto.objects.filter(codigo_sku=row['codigo_sku']).first()

                    if producto_existente:
                        if producto_existente.nombre.lower() != str(row['nombre']).lower():
                            errores.append(f"Fila {fila_num}: El código SKU '{row['codigo_sku']}' ya existe con un producto diferente ('{producto_existente.nombre}'). No se puede usar el mismo SKU para productos distintos.")
                            continue
                        
                        if producto_existente.categoria != categoria:
                            errores.append(f"Fila {fila_num}: El código SKU '{row['codigo_sku']}' existe pero con categoría diferente. SKU registrado: '{producto_existente.categoria.nombre}', Excel: '{categoria.nombre}'")
                            continue
                        
                        if producto_existente.subcategoria and subcategoria:
                            if producto_existente.subcategoria != subcategoria:
                                errores.append(f"Fila {fila_num}: El código SKU '{row['codigo_sku']}' existe pero con subcategoría diferente. SKU registrado: '{producto_existente.subcategoria.nombre}', Excel: '{subcategoria.nombre}'")
                                continue
                        
                        stock_a_sumar = int(row.get('stock', 0)) if not pd.isna(row.get('stock')) else 0
                        
                        if stock_a_sumar > 0:
                            stock_anterior = producto_existente.stock
                            producto_existente.stock += stock_a_sumar
                            producto_existente.save()
                            
                            MovimientoStock.objects.create(
                                producto=producto_existente,
                                tipo='ENTRADA',
                                cantidad=stock_a_sumar,
                                stock_anterior=stock_anterior,
                                stock_nuevo=producto_existente.stock,
                                motivo=f'Importación masiva desde Excel',
                                observaciones=f'Importado desde archivo Excel - Fila {fila_num}',
                                usuario=request.user if request.user.is_authenticated else None
                            )
                            
                            productos_actualizados += 1
                        
                        continue
                    
                    producto = Producto(
                        codigo_sku=str(row['codigo_sku']),
                        nombre=str(row['nombre']),
                        categoria=categoria,
                        subcategoria=subcategoria,
                        descripcion=str(row.get('descripcion', '')) if not pd.isna(row.get('descripcion')) else '',
                        precio=precio,
                        stock=int(row.get('stock', 0)) if not pd.isna(row.get('stock')) else 0,
                        stock_minimo=int(row.get('stock_minimo', 5)) if not pd.isna(row.get('stock_minimo')) else 5,
                        stock_critico=int(row.get('stock_critico', 2)) if not pd.isna(row.get('stock_critico')) else 2,
                    )
                    
                    productos_validos.append(producto)
                    
                except Exception as e:
                    errores.append(f"Fila {fila_num}: Error de formato en los datos. Revise que todas las columnas tengan el formato correcto.")
            
            productos_nuevos = 0
            if productos_validos:
                Producto.objects.bulk_create(productos_validos)
                productos_nuevos = len(productos_validos)

            if productos_nuevos > 0 and productos_actualizados > 0:
                messages.success(request, f'✓ Se importaron {productos_nuevos} productos nuevos y se actualizó el stock de {productos_actualizados} productos existentes')
            elif productos_nuevos > 0:
                messages.success(request, f'✓ Se importaron {productos_nuevos} productos nuevos exitosamente')
            elif productos_actualizados > 0:
                messages.success(request, f'✓ Se actualizó el stock de {productos_actualizados} productos existentes')
            
            if errores:
                request.session['errores_importacion'] = errores[:20]
                request.session.modified = True
            else:
                if 'errores_importacion' in request.session:
                    del request.session['errores_importacion']
                    request.session.modified = True
            
            return redirect('inventario:importar_productos')
        
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
            return redirect('inventario:importar_productos')


@solo_vendedor_o_admin
def descargar_plantilla(request):
    """Generar y descargar plantilla Excel para importacion - Vendedores y Admin"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Productos"
    
    headers = ['codigo_sku', 'nombre', 'categoria', 'subcategoria', 'descripcion', 'precio', 'stock', 'stock_minimo', 'stock_critico']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        
    ejemplos = [
        ['MC-0001', 'Starter Deck Digimon TCG', 'Decks', 'Digimon', 'Deck de inicio', 15990, 10, 5, 2],
        ['MC-0002', 'Sobre Pokemon Escarlata', 'Sobres', 'Pokemon', 'Sobre de expansión', 4500, 50, 10, 3],
        ['MC-0003', 'Figura Luffy Gear 5', 'Figuras', 'One Piece', 'Figura coleccionable', 45990, 5, 2, 1],
    ]
    
    for row_num, ejemplo in enumerate(ejemplos, 2):
        for col_num, value in enumerate(ejemplo, 1):
            ws.cell(row=row_num, column=col_num, value=value)
            
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_productos.xlsx'
    
    wb.save(response)
    return response


# ====================================
# VISTAS POS / VENTAS
# ====================================

@solo_vendedor_o_admin
def pos(request):
    """Vista principal del POS (Punto de Venta) - Solo vendedores y admin"""
    productos = Producto.objects.filter(activo=True, stock__gt=0).select_related('categoria', 'subcategoria')
    categorias = Categoria.objects.filter(activo=True)
    
    busqueda = request.GET.get('busqueda')
    categoria_filtro = request.GET.get('categoria')
    
    if busqueda:
        productos = productos.filter(
            Q(codigo_sku__icontains=busqueda) | Q(nombre__icontains=busqueda)
        )
    
    if categoria_filtro:
        productos = productos.filter(categoria_id=categoria_filtro)
    
    context = {
        'productos': productos,
        'categorias': categorias,
    }
    
    return render(request, 'inventario/pos.html', context)


@solo_vendedor_o_admin
def buscar_producto_ajax(request):
    """Búsqueda de productos vía AJAX para el POS - Solo vendedores y admin"""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'productos': []})
    
    productos = Producto.objects.filter(
        Q(codigo_sku__icontains=query) | Q(nombre__icontains=query),
        activo=True,
        stock__gt=0
    ).select_related('categoria', 'subcategoria')[:10]
    
    productos_data = []
    for p in productos:
        productos_data.append({
            'id': p.id,
            'codigo_sku': p.codigo_sku,
            'nombre': p.nombre,
            'precio': float(p.precio),
            'stock': p.stock,
            'imagen_url': p.get_imagen_url(),
            'categoria': p.categoria.nombre,
            'subcategoria': p.subcategoria.nombre if p.subcategoria else '',
        })
    
    return JsonResponse({'productos': productos_data})


@solo_vendedor_o_admin
def procesar_venta(request):
    """Procesar una venta desde el POS - Solo vendedores y admin"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            
            carrito = data.get('carrito', [])
            cliente_nombre = data.get('cliente_nombre', '').strip()
            
            if not carrito:
                return JsonResponse({
                    'success': False,
                    'error': 'El carrito está vacío'
                }, status=400)
            
            for item in carrito:
                producto = Producto.objects.get(id=item['producto_id'])
                if producto.stock < item['cantidad']:
                    return JsonResponse({
                        'success': False,
                        'error': f'Stock insuficiente para {producto.nombre}. Disponible: {producto.stock}'
                    }, status=400)
            
            venta = Venta.objects.create(
                cliente_nombre=cliente_nombre if cliente_nombre else None,
                usuario=request.user if request.user.is_authenticated else None
            )
            
            for item in carrito:
                producto = Producto.objects.get(id=item['producto_id'])
                
                DetalleVenta.objects.create(
                    venta=venta,
                    producto=producto,
                    cantidad=item['cantidad'],
                    precio_unitario=producto.precio
                )
                
                stock_anterior = producto.stock
                producto.stock -= item['cantidad']
                producto.save()
                
                MovimientoStock.objects.create(
                    producto=producto,
                    tipo='VENTA',
                    cantidad=item['cantidad'],
                    stock_anterior=stock_anterior,
                    stock_nuevo=producto.stock,
                    motivo=f'Venta {venta.folio}',
                    usuario=request.user if request.user.is_authenticated else None
                )
            
            venta.calcular_totales()
            
            return JsonResponse({
                'success': True,
                'venta_id': venta.id,
                'folio': venta.folio,
                'total': float(venta.total)
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'error': 'Método no permitido'
    }, status=405)


@solo_vendedor_o_admin
def lista_ventas(request):
    """Lista de ventas realizadas - Solo vendedores y admin"""
    ventas = Venta.objects.all().select_related('usuario').prefetch_related('detalles__producto')
    
    estado_filtro = request.GET.get('estado')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    busqueda = request.GET.get('busqueda')
    
    if estado_filtro:
        ventas = ventas.filter(estado=estado_filtro)
    
    if fecha_desde:
        ventas = ventas.filter(fecha_venta__date__gte=fecha_desde)
    
    if fecha_hasta:
        ventas = ventas.filter(fecha_venta__date__lte=fecha_hasta)
    
    if busqueda:
        ventas = ventas.filter(
            Q(folio__icontains=busqueda) | 
            Q(cliente_nombre__icontains=busqueda)
        )
    
    total_ventas = ventas.filter(estado='COMPLETADA').count()
    total_monto = sum(v.total for v in ventas.filter(estado='COMPLETADA'))
    ventas_anuladas = ventas.filter(estado='ANULADA').count()
    
    context = {
        'ventas': ventas,
        'total_ventas': total_ventas,
        'total_monto': total_monto,
        'ventas_anuladas': ventas_anuladas,
    }
    
    return render(request, 'inventario/lista_ventas.html', context)


@solo_vendedor_o_admin
def comprobante_venta(request, pk):
    """Generar comprobante de venta - Solo vendedores y admin"""
    venta = get_object_or_404(Venta, pk=pk)
    detalles = venta.detalles.all().select_related('producto')
    
    context = {
        'venta': venta,
        'detalles': detalles,
    }
    
    return render(request, 'inventario/comprobante_venta.html', context)


@solo_vendedor_o_admin
def anular_venta(request, pk):
    """Anular una venta y reponer el stock - Solo vendedores y admin"""
    venta = get_object_or_404(Venta, pk=pk)
    
    if request.method == 'POST':
        if venta.estado == 'ANULADA':
            messages.warning(request, f'La venta {venta.folio} ya está anulada')
            return redirect('inventario:lista_ventas')
        
        try:
            for detalle in venta.detalles.all():
                producto = detalle.producto
                stock_anterior = producto.stock
                producto.stock += detalle.cantidad
                producto.save()
                
                MovimientoStock.objects.create(
                    producto=producto,
                    tipo='ANULACION',
                    cantidad=detalle.cantidad,
                    stock_anterior=stock_anterior,
                    stock_nuevo=producto.stock,
                    motivo=f'Anulación de venta {venta.folio}',
                    observaciones=f'Se repone stock por anulación de venta',
                    usuario=request.user if request.user.is_authenticated else None
                )
            
            venta.estado = 'ANULADA'
            venta.save()
            
            messages.success(request, f'Venta {venta.folio} anulada exitosamente. Stock repuesto.')
            
        except Exception as e:
            messages.error(request, f'Error al anular venta: {str(e)}')
        
        return redirect('inventario:lista_ventas')
    return redirect('inventario:lista_ventas')